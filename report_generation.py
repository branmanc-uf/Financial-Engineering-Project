#ALL HERE ARE PENDING TESTING
#author: Antonio de Guzman

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from strategy_test import test_strategy
import matplotlib.pyplot as plt
import pandas as pd
from datetime import datetime, timedelta
from getting_options import black_scholes_dataframe

def generate_excel_report(symbol, df, output_file):
    """
    Generates an Excel report with strategy performance metrics and a candlestick chart using the new strategy function.

    Parameters:
    - symbol: Stock ticker (e.g., "SPY").
    - df: DataFrame containing stock data.
    - output_file: Path to save the Excel report.
    """
    # Run the strategy test and get performance metrics
    performance_metrics = test_strategy_from_df(
        df,
        initial_capital=100000,
        risk_per_trade=0.02,
        open_range_min=15,
        stop_loss_pct=0.02,
        take_profit_pct=0.04,
        expiration_days=30,
        sigma=0.2
    )

    if performance_metrics is None:
        print("⚠️ No data available or strategy test failed.")
        return

    # Extract metrics and trade log
    metrics = {k: v for k, v in performance_metrics.items() if k != 'DataFrame' and k != 'Trade Log'}
    trade_log = performance_metrics['Trade Log']

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Metrics"

    # Write performance metrics to the sheet
    ws.append(["Metric", "Value"])
    for key, value in metrics.items():
        ws.append([key, value])

    # Write trade log to a new sheet
    ws_trades = wb.create_sheet(title="Trade Log")
    for row in dataframe_to_rows(trade_log, index=False, header=True):
        ws_trades.append(row)

    # Generate and save the candlestick chart
    from generate_candlestick_df import plot_candlestick_with_indicators
    plot_candlestick_with_indicators(df, df)
    chart_path = "candlestick_chart.png"
    plt.savefig(chart_path, bbox_inches="tight")
    plt.close()

    # Embed the chart in the Excel report
    ws_chart = wb.create_sheet(title="Candlestick Chart")
    img = Image(chart_path)
    img.anchor = "A1"
    ws_chart.add_image(img)

    # Save the workbook
    wb.save(output_file)
    print(f"Excel report generated: {output_file}")

def generate_text_report(symbol, df, output_file):
    """
    Generates a text file summarizing the results of the strategy test.

    Parameters:
    - symbol: Stock ticker (e.g., "SPY").
    - df: DataFrame containing stock data.
    - output_file: Path to save the text report.
    """
    # Run the strategy test and get performance metrics
    performance_metrics = test_strategy_from_df(
        df,
        initial_capital=100000,
        risk_per_trade=0.02,
        open_range_min=15,
        stop_loss_pct=0.02,
        take_profit_pct=0.04,
        expiration_days=30,
        sigma=0.2
    )

    if performance_metrics is None:
        print("⚠️ No data available or strategy test failed.")
        return

    # Extract metrics and trade log
    metrics = {k: v for k, v in performance_metrics.items() if k != 'DataFrame' and k != 'Trade Log'}
    trade_log = performance_metrics['Trade Log']

    # Write performance metrics to a text file
    with open(output_file, "w") as file:
        file.write(f"Strategy Test Results for {symbol}\n")
        file.write("=" * 50 + "\n")
        for key, value in metrics.items():
            file.write(f"{key}: {value}\n")
        file.write("\nTrade Log:\n")
        file.write(trade_log.to_string(index=False))
    print(f"Text report generated: {output_file}")

# Example usage
df = pd.read_csv("df_2023_2025.csv")  # Replace with actual DataFrame loading
df['datetime'] = pd.to_datetime(df['datetime'])  # Ensure datetime column is parsed
generate_excel_report("SPY", df, "strategy_report.xlsx")
generate_text_report("SPY", df, "strategy_report.txt")
