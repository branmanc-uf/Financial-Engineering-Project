# ALL HERE ARE PENDING TESTING
# author: Antonio de Guzman

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from strategy_test_df import test_strategy_from_df
import matplotlib.pyplot as plt
import pandas as pd
from datetime import datetime, timedelta
from getting_options import black_scholes_dataframe

def generate_reports(symbol, df, output_file, strategy_params, image_paths=None):
    """
    Generates an Excel report with strategy performance metrics, a trade log, and optional images.

    Parameters:
    - symbol: Stock ticker (e.g., "SPY").
    - df: DataFrame containing stock data.
    - output_file: Path to save the Excel report.
    - strategy_params: Dictionary of parameters to pass to test_strategy_from_df.
    - image_paths: List of file paths to images to embed in the report.
    """

    performance_metrics = test_strategy_from_df(
        df,
        **strategy_params
    )

    if performance_metrics is None:
        print("⚠️ No data available or strategy test failed.")
        return

    # Extract metrics and trade log
    metrics = {k: v for k, v in performance_metrics.items() if k != 'DataFrame' and k != 'Trade Log'}
    trade_log = performance_metrics['Trade Log']

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Metrics"

    # Write performance metrics to the sheet
    ws.append(["Metric", "Value"])
    for key, value in metrics.items():
        ws.append([key, value])

    # Add images to the "Performance Metrics" sheet
    if image_paths:
        current_row = ws.max_row + 2  # Leave a blank row after the metrics
        for image_path in image_paths:
            try:
                img = Image(image_path)
                img.anchor = f"A{current_row}"  # Place the image starting at column A
                ws.add_image(img)
                current_row += 70  # Leave space for the image (adjust as needed)
            except Exception as e:
                print(f"⚠️ Could not add image {image_path}: {e}")

    # Write trade log to a new sheet
    ws_trades = wb.create_sheet(title="Trade Log")

    # Strip timezone from datetime before exporting to Excel
    trade_log_export = trade_log.copy()

    # Strip timezone from any datetime-like columns
    for col in trade_log_export.columns:
        if pd.api.types.is_datetime64_any_dtype(trade_log_export[col]):
            trade_log_export[col] = trade_log_export[col].dt.tz_localize(None)


    for row in dataframe_to_rows(trade_log_export, index=False, header=True):
        ws_trades.append(row)

    # Save the workbook
    wb.save(output_file)
    print(f"Excel report generated: {output_file}")
    
    # Write performance metrics to a text file
    text_output_file = output_file.replace(".xlsx", ".txt")
    with open(text_output_file, "w") as file:
        file.write(f"Strategy Test Results for {symbol}\n")
        file.write("=" * 50 + "\n")
        for key, value in metrics.items():
            file.write(f"{key}: {value}\n")
        file.write("\nTrade Log:\n")
        file.write(trade_log.to_string(index=False))
    print(f"Text report generated: {text_output_file}")
